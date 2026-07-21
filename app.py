import streamlit as st
import streamlit.components.v1 as components
import requests, json, time, hashlib, base64, os
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
st.set_page_config(page_title="Controle Frotas", page_icon="🌾",
                   layout="wide", initial_sidebar_state="collapsed")

# ─────────────────────────────────────────────
# LOGIN (2 perfis: admin / visualizador)
# ─────────────────────────────────────────────
# Usuários ficam nos secrets, em [usuarios], como:
#   [usuarios.joao]
#   senha_sha256 = "..."   # gerado com gerar_senha.py
#   perfil = "admin"       # ou "visualizador"
USUARIOS = st.secrets.get("usuarios", {})

def _hash(senha: str) -> str:
    return hashlib.sha256(senha.encode("utf-8")).hexdigest()

def _autenticar(usuario: str, senha: str):
    u = USUARIOS.get(usuario)
    if not u:
        return None
    if _hash(senha) != u.get("senha_sha256", ""):
        return None
    return u.get("perfil", "visualizador")

if "auth_usuario" not in st.session_state:
    st.session_state["auth_usuario"] = None
    st.session_state["auth_perfil"] = None

if not st.session_state["auth_usuario"]:
    st.markdown("### 🌾 Controle Frotas — Teston / Metalcana")
    st.caption("Faça login para continuar.")
    _c1, _c2, _c3 = st.columns([1, 1, 1])
    with _c2:
        with st.form("form_login"):
            _login_user = st.text_input("Usuário")
            _login_pass = st.text_input("Senha", type="password")
            _login_btn  = st.form_submit_button("Entrar", type="primary", use_container_width=True)
        if _login_btn:
            if not USUARIOS:
                st.error("Nenhum usuário configurado em st.secrets['usuarios']. Avise o admin.")
            else:
                _perfil = _autenticar(_login_user.strip(), _login_pass)
                if _perfil:
                    st.session_state["auth_usuario"] = _login_user.strip()
                    st.session_state["auth_perfil"]  = _perfil
                    st.rerun()
                else:
                    st.error("Usuário ou senha incorretos.")
    st.stop()

PERFIL = st.session_state["auth_perfil"]
PODE_EDITAR = PERFIL == "admin"

# ── CONFIGURAÇÃO ─────────────────────────────
TENANT_ID     = st.secrets["TENANT_ID"]
CLIENT_ID     = st.secrets["CLIENT_ID"]
CLIENT_SECRET = st.secrets["CLIENT_SECRET"]
SITE_HOST     = st.secrets.get("SITE_HOST", "metalcana.sharepoint.com")
SITE_NAME     = st.secrets.get("SITE_NAME", "AppKanbanFrotas")
LISTA_CC      = st.secrets.get("LISTA_CC",      "KanbanCC")
LISTA_FRENTES = st.secrets.get("LISTA_FRENTES", "KanbanFrentes")
LISTA_FROTAS  = st.secrets.get("LISTA_FROTAS",  "KanbanFrotas")
LISTA_ENTREGA = st.secrets.get("LISTA_ENTREGA_FUTURA", "KanbanEntregaFutura")

TIPOS = ["Colhedora","Transbordo","Trator","Caminhão","Veículo","Implemento","Apoio","Área de Vivência","Outro"]
# Ícones SVG customizados (inline, sem dependência externa) — herdam a cor do texto
SVG_COLHEDORA  = ("<svg viewBox='0 0 24 24' width='15' height='15' style='vertical-align:-2px' "
                  "xmlns='http://www.w3.org/2000/svg'><g fill='currentColor'>"
                  "<path d='M2 16.2l3-6.2v6.2z'/>"
                  "<rect x='5.2' y='10.3' width='10' height='5.9' rx='.8'/>"
                  "<path d='M6.2 10.3V6.1c0-.7.6-1.3 1.3-1.3h3.2c.7 0 1.3.6 1.3 1.3v4.2z'/>"
                  "<path d='M14.3 11.2l6.5-6.5 1.9 1.9-6.5 6.5z'/>"
                  "<circle cx='21.9' cy='5.7' r='1.6'/>"
                  "<circle cx='7.6' cy='18.1' r='2.3'/><circle cx='13.6' cy='18.1' r='2.3'/>"
                  "</g></svg>")
SVG_IMPLEMENTO = ("<svg viewBox='0 0 24 24' width='15' height='15' style='vertical-align:-2px' "
                  "xmlns='http://www.w3.org/2000/svg'><g fill='currentColor'>"
                  "<path d='M12 2.2l4.8 4.8H7.2z'/>"
                  "<rect x='1.8' y='7.6' width='20.4' height='2.4' rx='1'/>"
                  "<rect x='4.7' y='9.8' width='1.4' height='3.4'/><rect x='9.2' y='9.8' width='1.4' height='3.4'/>"
                  "<rect x='13.7' y='9.8' width='1.4' height='3.4'/><rect x='18.2' y='9.8' width='1.4' height='3.4'/>"
                  "<circle cx='5.4' cy='16' r='3'/><circle cx='9.9' cy='16' r='3'/>"
                  "<circle cx='14.4' cy='16' r='3'/><circle cx='18.9' cy='16' r='3'/>"
                  "</g></svg>")
SVG_APOIO      = ("<svg viewBox='0 0 24 24' width='15' height='15' style='vertical-align:-2px' "
                  "xmlns='http://www.w3.org/2000/svg'><g fill='currentColor'>"
                  "<path d='M2 10V7.5C2 5 4 3 6.5 3h11C20 3 22 5 22 7.5V10z'/>"
                  "<path d='M2 11.5h7.5V14h5v-2.5H22V21H2z'/>"
                  "<rect x='10' y='8.5' width='4' height='4.6' rx='.8'/>"
                  "</g></svg>")
def _icone_png(nome, fallback, px=16):
    """Procura <nome>.png na raiz do projeto ou na pasta icones/;
    senão, usa o fallback (SVG desenhado). Ideal: PNGs de até 64x64 px."""
    raiz = os.path.dirname(os.path.abspath(__file__))
    for caminho in (os.path.join(raiz, f"{nome}.png"),
                    os.path.join(raiz, "icones", f"{nome}.png")):
        try:
            with open(caminho, "rb") as fp:
                b64 = base64.b64encode(fp.read()).decode()
            return (f"<img src='data:image/png;base64,{b64}' width='{px}' height='{px}' "
                    f"style='vertical-align:-3px;object-fit:contain'/>")
        except Exception:
            continue
    return fallback

TIPO_EMOJI = {"Colhedora": _icone_png("colhedora", SVG_COLHEDORA),
               "Transbordo":"🚛","Trator":"🚜","Caminhão":"🚚",
               "Veículo":"🚗",
               "Implemento": _icone_png("implemento", SVG_IMPLEMENTO),
               "Apoio": _icone_png("apoio", SVG_APOIO),
               "Área de Vivência":"🏠","Outro":"📦"}
CORES_CC = ["#1D9E75","#378ADD","#BA7517","#D85A30","#7F77DD","#D4537E",
            "#639922","#0F6E56","#185FA5","#854F0B","#993C1D","#534AB7",
            "#3B6D11","#A32D2D","#5b21b6","#9d174d","#166534","#374151"]
TIPO_COR = {
    "Colhedora":  ["#d1fae5","#065f46"], "Transbordo": ["#dbeafe","#1e40af"],
    "Trator":     ["#fef3c7","#92400e"], "Caminhão":   ["#ede9fe","#5b21b6"],
    "Veículo":    ["#fce7f3","#9d174d"], "Implemento": ["#f0fdf4","#166534"],
    "Apoio":      ["#fee2e2","#991b1b"], "Área de Vivência": ["#e0f2fe","#075985"],
    "Outro":      ["#f3f4f6","#374151"],
}

# ─────────────────────────────────────────────
# GRAPH API
# ─────────────────────────────────────────────
@st.cache_data(ttl=3500)
def get_token():
    r = requests.post(
        f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token",
        data={"grant_type":"client_credentials","client_id":CLIENT_ID,
              "client_secret":CLIENT_SECRET,"scope":"https://graph.microsoft.com/.default"})
    r.raise_for_status()
    return r.json()["access_token"]

@st.cache_data(ttl=3500)
def get_site_id():
    token = get_token()
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{SITE_HOST}:/sites/{SITE_NAME}",
        headers={"Authorization": f"Bearer {token}"})
    r.raise_for_status()
    return r.json()["id"]

def hdrs():
    return {"Authorization": f"Bearer {get_token()}", "Content-Type": "application/json"}

def lista_items(lista_nome):
    site_id = get_site_id()
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{lista_nome}/items?expand=fields&$top=2000"
    r = requests.get(url, headers=hdrs())
    r.raise_for_status()
    return r.json().get("value", [])

def patch_item(lista_nome, item_id, fields):
    site_id = get_site_id()
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{lista_nome}/items/{item_id}/fields"
    r = requests.patch(url, headers=hdrs(), json=fields)
    r.raise_for_status()
    return r.json()

def criar_item(lista_nome, fields):
    site_id = get_site_id()
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{lista_nome}/items"
    r = requests.post(url, headers=hdrs(), json={"fields": fields})
    r.raise_for_status()
    return r.json()

# ─────────────────────────────────────────────
# CARREGAR DADOS
# ─────────────────────────────────────────────
@st.cache_data(ttl=30)
def carregar_dados():
    ccs_raw     = lista_items(LISTA_CC)
    frentes_raw = lista_items(LISTA_FRENTES)
    frotas_raw  = lista_items(LISTA_FROTAS)

    ccs = sorted([{
        "id":    item["id"],
        "nome":  item["fields"].get("Title",""),
        "ordem": int(item["fields"].get("Ordem", 99)),
    } for item in ccs_raw], key=lambda x: x["ordem"])

    frentes = [{"id": item["id"], "nome": item["fields"].get("Title","")}
               for item in frentes_raw]

    frotas = [{
        "id":          item["id"],
        "nome":        item["fields"].get("Title",""),
        "tipo":        item["fields"].get("Tipo","Outro"),
        "chassi":      item["fields"].get("Chassi",""),
        "ano":         item["fields"].get("Ano",""),
        "obs":         item["fields"].get("Obs",""),
        "cc_nome":     item["fields"].get("CCNome",""),
        "frente_nome": item["fields"].get("FrenteNome",""),
        "status":      item["fields"].get("Status","Ativo"),
        "data_venda":  item["fields"].get("DataVenda",""),
        "valor_venda": item["fields"].get("ValorVenda",""),
    } for item in frotas_raw]

    return ccs, frentes, frotas

@st.cache_data(ttl=30)
def carregar_entregas():
    entregas_raw = lista_items(LISTA_ENTREGA)
    return [{
        "id":            item["id"],
        "nome":          item["fields"].get("Title",""),
        "frota_id":      item["fields"].get("FrotaItemId",""),
        "chassi":        item["fields"].get("Chassi",""),
        "tipo":          item["fields"].get("Tipo",""),
        "cc_origem":     item["fields"].get("CCOrigem",""),
        "frente_origem": item["fields"].get("FrenteOrigem",""),
        "motivo":        item["fields"].get("Motivo",""),
        "destino":       item["fields"].get("Destino",""),
        "status":        item["fields"].get("Status","Pendente"),
        "data_cadastro": item["fields"].get("DataCadastro",""),
        "data_entrega":  item["fields"].get("DataEntrega",""),
        "obs_baixa":     item["fields"].get("ObsBaixa",""),
    } for item in entregas_raw]

def invalidar():
    carregar_dados.clear()
    carregar_entregas.clear()

# ─────────────────────────────────────────────
# AÇÕES
# ─────────────────────────────────────────────
def mover_cc(frota_id, novo_cc):
    patch_item(LISTA_FROTAS, frota_id, {"CCNome": novo_cc, "FrenteNome": ""})
    invalidar()

def mover_frente(frota_id, nova_frente):
    patch_item(LISTA_FROTAS, frota_id, {"FrenteNome": nova_frente})
    invalidar()

def vender_frota(frota_id, data_venda, valor_venda, obs_venda):
    patch_item(LISTA_FROTAS, frota_id, {
        "Status": "Vendido", "CCNome": "", "FrenteNome": "",
        "DataVenda": data_venda, "ValorVenda": valor_venda,
        "Obs": obs_venda,
    })
    invalidar()

def reativar_frota(frota_id):
    patch_item(LISTA_FROTAS, frota_id, {
        "Status": "Ativo", "DataVenda": "", "ValorVenda": "",
    })
    invalidar()

def registrar_entrega_futura(frota, motivo, destino):
    """Cria um registro de 'entrega futura' vinculado a uma frota existente.
    Não altera CCNome/FrenteNome da frota — é só um aviso à parte."""
    criar_item(LISTA_ENTREGA, {
        "Title":        frota["nome"],
        "FrotaItemId":  frota["id"],
        "Chassi":       frota.get("chassi",""),
        "Tipo":         frota.get("tipo",""),
        "CCOrigem":     frota.get("cc_nome",""),
        "FrenteOrigem": frota.get("frente_nome",""),
        "Motivo":       motivo,
        "Destino":      destino,
        "Status":       "Pendente",
        "DataCadastro": datetime.today().strftime("%Y-%m-%d"),
        "DataEntrega":  "",
        "ObsBaixa":     "",
    })
    invalidar()

def marcar_entrega_concluida(entrega_id, data_entrega, obs_baixa):
    patch_item(LISTA_ENTREGA, entrega_id, {
        "Status": "Entregue", "DataEntrega": data_entrega, "ObsBaixa": obs_baixa,
    })
    invalidar()

def cancelar_entrega_futura(entrega_id):
    """Remove um aviso de entrega futura cadastrado por engano."""
    patch_item(LISTA_ENTREGA, entrega_id, {"Status": "Cancelado"})
    invalidar()

# ─────────────────────────────────────────────
# CARREGAR
# ─────────────────────────────────────────────
try:
    ccs, frentes, frotas = carregar_dados()
    entregas = carregar_entregas()
except Exception as e:
    st.error(f"Erro ao conectar ao SharePoint: {e}")
    st.stop()

# Separa ativos e vendidos
frotas_ativas  = [f for f in frotas if f.get("status","Ativo") != "Vendido"]
frotas_vendidas = [f for f in frotas if f.get("status","Ativo") == "Vendido"]

# Separa entregas futuras pendentes e concluídas
entregas_pendentes = [e for e in entregas if e.get("status","Pendente") == "Pendente"]
entregas_concluidas = [e for e in entregas if e.get("status") == "Entregue"]
ids_frota_com_entrega_pendente = {e["frota_id"] for e in entregas_pendentes if e.get("frota_id")}

# ─────────────────────────────────────────────
# PROCESSAR AÇÕES (query params)
# ─────────────────────────────────────────────
# As ações do board (mover, salvar obs, vender) agora chegam via componente
# bidirecional (postMessage) — sem recarregar a página e sem derrubar o login.
# O visual/JS do board fica em board_component/index.html.
_BOARD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "board_component")
kanban_board = components.declare_component("kanban_board", path=_BOARD_DIR)

# ─────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
#MainMenu, footer, header {visibility: hidden;}
.block-container {padding: 1rem 1.5rem !important;}
.stTabs [data-baseweb="tab-list"] {gap: 8px;}
.stTabs [data-baseweb="tab"] {padding: 6px 16px; border-radius: 8px 8px 0 0;}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# ABAS PRINCIPAIS
# ─────────────────────────────────────────────
st.markdown("### 🌾 Controle Frotas — Teston / Metalcana")
_hc1, _hc2 = st.columns([5, 1])
with _hc1:
    _badge = "🔑 admin" if PODE_EDITAR else "👁️ visualizador"
    st.caption(f"Usuário: **{st.session_state['auth_usuario']}** · Perfil: {_badge}")
with _hc2:
    if st.button("Sair", use_container_width=True):
        st.session_state["auth_usuario"] = None
        st.session_state["auth_perfil"] = None
        st.rerun()
aba_board, aba_nova, aba_entrega, aba_vendidos, aba_export = st.tabs([
    "🗂️ Board Controle",
    "➕ Nova Frota",
    f"🚚 Entrega Futura ({len(entregas_pendentes)})",
    "📋 Histórico de Baixas",
    "📤 Exportar Relatório",
])

# ══════════════════════════════════════════════
# ABA 1 — BOARD KANBAN
# ══════════════════════════════════════════════
with aba_board:

    # ── FILTROS ──────────────────────────────
    f1, f2, f3 = st.columns([2, 2, 2])
    with f1:
        filtro_tipo = st.multiselect("🔧 Tipo", TIPOS, default=[], placeholder="Todos os tipos")
    with f2:
        busca = st.text_input("🔍 Buscar frota", placeholder="Número ou nome...")
    with f3:
        nomes_frente_opc = ["— Todas as frentes —"] + [fr["nome"] for fr in frentes]
        filtro_frente = st.selectbox("⚡ Buscar frente", nomes_frente_opc,
                                     help="Filtra todos os equipamentos da frente selecionada")

    f4, f5, f6 = st.columns([2, 2, 2])
    with f4:
        nomes_cc_opc = ["— Todos os CCs —"] + [c["nome"] for c in ccs]
        filtro_origem_idx = 0
        # Auto-preenche origem quando há busca com resultado único
        if busca:
            matches = [f for f in frotas_ativas if busca.lower() in f["nome"].lower()]
            if len(matches) == 1 and matches[0]["cc_nome"]:
                cc_match = matches[0]["cc_nome"]
                if cc_match in nomes_cc_opc:
                    filtro_origem_idx = nomes_cc_opc.index(cc_match)
        filtro_cc_origem = st.selectbox("📍 CC de origem", nomes_cc_opc,
                                        index=filtro_origem_idx,
                                        help="Preenchido automaticamente ao buscar uma frota")
    with f5:
        nomes_cc_dest = ["— Sem filtro destino —"] + [c["nome"] for c in ccs] + ["📦 Sem alocação"]
        filtro_cc_destino = st.selectbox("🎯 CC de destino", nomes_cc_dest)
    with f6:
        busca_cc = st.text_input("🔍 Buscar centro de custo", placeholder="Nome do CC...",
                                 help="Mostra só as colunas de CC que contêm o texto digitado")

    # ── LÓGICA DO BOARD ───────────────────────
    ccs_todos = ccs + [{"id":"__sem__","nome":"📦 Sem alocação","ordem":999}]
    cc_origem_ativo  = filtro_cc_origem  != "— Todos os CCs —"
    cc_destino_ativo = filtro_cc_destino != "— Sem filtro destino —"

    def frota_vis(f):
        if filtro_tipo and f["tipo"] not in filtro_tipo: return False
        if busca and busca.lower() not in f["nome"].lower(): return False
        if filtro_frente != "— Todas as frentes —" and f.get("frente_nome","") != filtro_frente: return False
        return True

    # Quais colunas mostrar
    frente_ativa = filtro_frente != "— Todas as frentes —"
    if cc_origem_ativo and cc_destino_ativo:
        nomes_mostrar = list(dict.fromkeys([filtro_cc_origem, filtro_cc_destino]))
        ccs_board = [c for c in ccs_todos if c["nome"] in nomes_mostrar]
        ccs_board.sort(key=lambda c: 0 if c["nome"] == filtro_cc_origem else 1)
    elif cc_origem_ativo:
        ccs_board = [c for c in ccs_todos if c["nome"] == filtro_cc_origem]
    elif frente_ativa:
        # Mostra só CCs que têm frotas dessa frente
        ccs_com_frente = {f["cc_nome"] if f["cc_nome"] else "📦 Sem alocação"
                          for f in frotas_ativas if f.get("frente_nome","") == filtro_frente}
        ccs_board = [c for c in ccs_todos if c["nome"] in ccs_com_frente]
    else:
        ccs_board = ccs_todos

    # Filtro por nome de centro de custo
    if busca_cc:
        ccs_board = [c for c in ccs_board if busca_cc.lower() in c["nome"].lower()]

    # IDs destacados (frotas que batem na busca/origem)
    def frota_destacada(f):
        if not frota_vis(f): return False
        if cc_origem_ativo:
            cc_f = f["cc_nome"] if f["cc_nome"] else "📦 Sem alocação"
            if cc_f != filtro_cc_origem: return False
        return True

    frotas_board   = [f for f in frotas_ativas if frota_vis(f)]
    ids_destacados = {f["id"] for f in frotas_ativas if frota_destacada(f)} if (cc_origem_ativo or busca or frente_ativa) else set()

    nomes_cc = [c["nome"] for c in ccs_todos]
    nomes_fr = ["— Sem frente —"] + [fr["nome"] for fr in frentes]

    frotas_por_cc = {
        cc["nome"]: [f for f in frotas_board
                     if f["cc_nome"] == (cc["nome"] if cc["id"] != "__sem__" else "")]
        for cc in ccs_board
    }

    board_data = {
        "pode_editar": PODE_EDITAR,
        "ccs": [{
            "nome":     c["nome"],
            "cor":      CORES_CC[i % len(CORES_CC)],
            "destaque": c["nome"] == filtro_cc_origem  if cc_origem_ativo  else False,
            "destino":  c["nome"] == filtro_cc_destino if cc_destino_ativo else False,
        } for i,c in enumerate(ccs_todos) if c in ccs_board],
        "ids_destacados": list(ids_destacados),
        "frotas_por_cc": {
            cc["nome"]: [{
                "id":        f["id"],
                "nome":      f["nome"],
                "tipo":      f["tipo"],
                "chassi":    f["chassi"],
                "ano":       f["ano"],
                "frente":    f["frente_nome"],
                "obs":       f.get("obs",""),
                "cor_bg":    TIPO_COR.get(f["tipo"],["#f3f4f6","#374151"])[0],
                "cor_tx":    TIPO_COR.get(f["tipo"],["#f3f4f6","#374151"])[1],
                "destacado": f["id"] in ids_destacados,
            } for f in frotas_por_cc[cc["nome"]]]
            for cc in ccs_board
        },
        "nomes_cc": nomes_cc,
        "nomes_fr": nomes_fr,
        "tipos_ordem": TIPOS,
        "tipo_emoji": TIPO_EMOJI,
    }

    # ── MÉTRICAS ─────────────────────────────
    total    = len(frotas_board)
    alocadas = sum(1 for f in frotas_board if f["cc_nome"])
    colh     = sum(1 for f in frotas_board if f["tipo"] == "Colhedora")
    sem_fr   = sum(1 for f in frotas_board if f["cc_nome"] and not f["frente_nome"])

    m1,m2,m3,m4 = st.columns(4)
    for col, val, lbl, cor in [
        (m1, total,    "Frotas ativas",      "#111"),
        (m2, alocadas, "Alocadas",           "#065f46"),
        (m3, colh,     "Colhedoras",         "#1D9E75"),
        (m4, sem_fr,   "Alocadas s/ frente", "#dc2626" if sem_fr else "#111"),
    ]:
        col.markdown(
            f'<div style="background:#f9fafb;border:1px solid #e5e7eb;border-radius:8px;'
            f'padding:8px 14px;text-align:center;margin-bottom:8px">'
            f'<div style="font-size:24px;font-weight:700;color:{cor}">{val}</div>'
            f'<div style="font-size:11px;color:#6b7280">{lbl}</div></div>',
            unsafe_allow_html=True)

    # ── (LEGADO — NÃO USADO) O board real agora está em board_component/index.html.
    # Este bloco antigo foi mantido apenas como referência e pode ser apagado.
    board_json = json.dumps(board_data, ensure_ascii=False)

    BOARD_HTML = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
*{{box-sizing:border-box;margin:0;padding:0;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;}}
body{{background:transparent;overflow-x:auto;}}
#board{{display:flex;gap:12px;padding:4px 2px 16px;align-items:flex-start;min-height:480px;}}
.column{{min-width:360px;flex:1 1 360px;max-width:50%;background:#f8fafc;border-radius:10px;border:1px solid #e2e8f0;overflow:hidden;}}
.col-header{{padding:9px 12px;display:flex;align-items:center;justify-content:space-between;color:#fff;font-size:12px;font-weight:700;border-radius:10px 10px 0 0;}}
.col-header.col-origem{{outline:3px solid #fbbf24;outline-offset:-2px;}}
.col-header.col-destino{{outline:3px solid #34d399;outline-offset:-2px;}}
.col-badge{{background:rgba(255,255,255,.28);border-radius:20px;padding:1px 8px;font-size:11px;}}
.col-label{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.05em;padding:2px 8px;text-align:center;}}
.col-label.origem{{background:#fef3c7;color:#92400e;}}
.col-label.destino{{background:#d1fae5;color:#065f46;}}
.cards-list{{min-height:60px;padding:8px;display:flex;flex-direction:column;gap:6px;transition:background .15s;}}
.cards-list.drag-over{{background:#e0f2fe;border-radius:0 0 8px 8px;}}
.card{{background:#fff;border:1px solid #e2e8f0;border-radius:8px;padding:8px 10px;cursor:grab;user-select:none;border-left:4px solid #1D9E75;transition:box-shadow .15s,transform .1s,opacity .15s;}}
.card.somente-leitura{{cursor:default;}}
.card:hover{{box-shadow:0 2px 8px rgba(0,0,0,.10);}}
.card:active{{cursor:grabbing;transform:scale(.97);}}
.card.dragging{{opacity:.35;}}
.card.drag-ghost{{box-shadow:0 8px 24px rgba(0,0,0,.18);transform:rotate(2deg) scale(1.03);opacity:.95;pointer-events:none;position:fixed;z-index:9999;width:340px;}}
.card.destacado{{border-left-width:5px;box-shadow:0 0 0 2px #fbbf24;background:#fffbeb;}}
.card.nao-destacado{{opacity:.25;filter:grayscale(.6);pointer-events:none;}}
.card-name{{font-size:12px;font-weight:600;color:#111;margin-bottom:2px;}}
.card-sub{{font-size:10px;color:#6b7280;margin-bottom:4px;}}
.card-tag{{display:inline-block;padding:1px 7px;border-radius:20px;font-size:10px;font-weight:600;margin-right:4px;}}
.card-frente{{font-size:10px;color:#374151;margin-top:4px;padding:2px 6px;background:#f1f5f9;border-radius:4px;display:inline-flex;align-items:center;gap:4px;cursor:pointer;}}
.card-frente:hover{{background:#dbeafe;color:#1e40af;}}
.tipo-header{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.05em;color:#475569;background:#e2e8f0;border-radius:5px;padding:3px 8px;margin-top:2px;}}
.card-obs-btn{{font-size:10px;color:#6b7280;margin-top:5px;padding:2px 6px;background:#f8fafc;border:1px dashed #d1d5db;border-radius:4px;cursor:pointer;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}}
.card-obs-btn:hover{{background:#eef2ff;color:#3730a3;}}
.card-obs-box textarea{{width:100%;min-height:54px;font-size:11px;padding:6px;border:1px solid #d1d5db;border-radius:6px;resize:vertical;margin-top:5px;font-family:inherit;color:#111;background:#fff;}}
.card-obs-save{{margin-top:4px;width:100%;padding:4px 0;font-size:11px;border-radius:6px;border:1px solid #0F6E56;background:#1D9E75;color:#fff;cursor:pointer;font-weight:600;}}
.card-obs-save:hover{{background:#0F6E56;}}
.empty-hint{{text-align:center;font-size:11px;color:#9ca3af;padding:14px 6px;border:1px dashed #d1d5db;border-radius:6px;}}
.modal-bg{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:10000;align-items:center;justify-content:center;}}
.modal-bg.open{{display:flex;}}
.modal{{background:#fff;border-radius:12px;padding:20px;width:340px;box-shadow:0 20px 60px rgba(0,0,0,.25);}}
.modal h3{{font-size:15px;font-weight:600;margin-bottom:6px;color:#111;}}
.modal .sub{{font-size:12px;color:#6b7280;margin-bottom:16px;}}
.modal select{{width:100%;padding:8px 10px;border:1px solid #d1d5db;border-radius:8px;font-size:13px;margin-bottom:14px;background:#fff;color:#111;}}
.modal-btns{{display:flex;gap:8px;justify-content:flex-end;}}
.btn{{padding:7px 16px;border-radius:8px;font-size:13px;cursor:pointer;border:1px solid #d1d5db;background:#fff;color:#111;}}
.btn:hover{{background:#f1f5f9;}}
.btn.primary{{background:#1D9E75;color:#fff;border-color:#0F6E56;}}
.btn.primary:hover{{background:#0F6E56;}}
.btn.danger{{background:#fee2e2;color:#991b1b;border-color:#fca5a5;}}
.btn.danger:hover{{background:#fecaca;}}
</style></head><body>
<div id="board"></div>
<div class="modal-bg" id="modalBg">
  <div class="modal">
    <h3 id="modalNome"></h3>
    <div class="sub" id="modalSub"></div>
    <select id="modalSelect"></select>
    <div class="modal-btns">
      <button class="btn danger" onclick="marcarVendido()">🪦 Marcar vendido</button>
      <button class="btn" onclick="fecharModal()">Cancelar</button>
      <button class="btn primary" onclick="salvarFrente()">Salvar frente</button>
    </div>
  </div>
</div>
<script>
const DATA = {board_json};
const nomesCc = DATA.nomes_cc;
const nomesFr = DATA.nomes_fr;
const PODE_EDITAR = DATA.pode_editar;
let dragFrotaId=null,dragFrotaNome=null,dragEl=null,ghostEl=null,modalFrotaId=null;

function render(){{
  const board=document.getElementById('board');
  board.innerHTML='';
  DATA.ccs.forEach(cc=>{{
    const cards=DATA.frotas_por_cc[cc.nome]||[];
    const col=document.createElement('div');
    col.className='column';
    const origemCls =cc.destaque?' col-origem':'';
    const destinoCls=cc.destino?' col-destino':'';
    const labelO=cc.destaque?'<div class="col-label origem">📍 Origem</div>':'';
    const labelD=cc.destino ?'<div class="col-label destino">🎯 Destino</div>':'';
    col.innerHTML=`
      <div class="col-header${{origemCls}}${{destinoCls}}" style="background:${{cc.cor}}">
        <span style="white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:290px" title="${{cc.nome}}">${{cc.nome}}</span>
        <span class="col-badge">${{cards.length}}</span>
      </div>${{labelO}}${{labelD}}
      <div class="cards-list" data-cc="${{cc.nome}}" id="list-${{cc.nome.replace(/[^a-z0-9]/gi,'_')}}"></div>`;
    board.appendChild(col);
    const list=col.querySelector('.cards-list');
    setupDrop(list,cc.nome);
    if(cards.length===0){{list.innerHTML='<div class="empty-hint">Sem frotas</div>';}}
    else{{
      const temDest=DATA.ids_destacados&&DATA.ids_destacados.length>0;
      const grupos={{}};
      cards.forEach(f=>{{(grupos[f.tipo]=grupos[f.tipo]||[]).push(f);}});
      const ordem=DATA.tipos_ordem.concat(Object.keys(grupos).filter(t=>!DATA.tipos_ordem.includes(t)));
      ordem.forEach(t=>{{
        const g=grupos[t];if(!g||!g.length)return;
        const gh=document.createElement('div');
        gh.className='tipo-header';
        gh.innerHTML=`${{DATA.tipo_emoji[t]||'📦'}} ${{t}} (${{g.length}})`;
        list.appendChild(gh);
        g.forEach(f=>list.appendChild(criarCard(f,cc.cor,temDest)));
      }});
    }}
  }});
}}

function criarCard(frota,corCC,temDest){{
  const el=document.createElement('div');
  let cls='card';
  if(temDest) cls+=frota.destacado?' destacado':' nao-destacado';
  if(!PODE_EDITAR) cls+=' somente-leitura';
  el.className=cls;
  el.dataset.id=frota.id;
  el.style.borderLeftColor=corCC;
  const sub=[frota.chassi,frota.ano].filter(Boolean).join(' | ');
  const frLabel=frota.frente||'+ frente';
  const frenteAttr=PODE_EDITAR?`onclick="abrirModal(event,'${{frota.id}}','${{frota.nome.replace(/'/g,"\\'")}}','${{(frota.frente||'').replace(/'/g,"\\'")}}')"`:'';
  el.innerHTML=`
    <div class="card-name">${{DATA.tipo_emoji[frota.tipo]||'📦'}} ${{frota.nome}}</div>
    ${{sub?`<div class="card-sub">${{sub}}</div>`:''}}
    <span class="card-tag" style="background:${{frota.cor_bg}};color:${{frota.cor_tx}}">${{frota.tipo}}</span>
    <div class="card-frente" ${{frenteAttr}} style="${{PODE_EDITAR?'':'cursor:default;opacity:.7'}}">⚡ ${{frLabel}}</div>`;
  // Observação no card
  const obsBtn=document.createElement('div');
  obsBtn.className='card-obs-btn';
  obsBtn.textContent=frota.obs?'📝 '+frota.obs:(PODE_EDITAR?'📝 + observação':'');
  if(frota.obs)obsBtn.title=frota.obs;
  if(frota.obs||PODE_EDITAR)el.appendChild(obsBtn);
  if(PODE_EDITAR){{
    const obsBox=document.createElement('div');
    obsBox.className='card-obs-box';
    obsBox.style.display='none';
    const ta=document.createElement('textarea');
    ta.value=frota.obs||'';
    ta.placeholder='Escreva a observação...';
    ta.addEventListener('mousedown',ev=>{{ev.stopPropagation();el.draggable=false;}});
    ta.addEventListener('blur',()=>{{el.draggable=true;}});
    const sv=document.createElement('button');
    sv.className='card-obs-save';
    sv.textContent='💾 Salvar observação';
    sv.onclick=ev=>{{ev.stopPropagation();salvarObs(frota.id,ta.value);}};
    obsBox.appendChild(ta);obsBox.appendChild(sv);
    el.appendChild(obsBox);
    obsBtn.onclick=ev=>{{ev.stopPropagation();
      obsBox.style.display=obsBox.style.display==='none'?'block':'none';
      if(obsBox.style.display==='block')ta.focus();}};
  }}
  el.draggable=PODE_EDITAR;
  if(!PODE_EDITAR){{ el.style.cursor='default'; return el; }}
  el.addEventListener('dragstart',e=>{{
    dragFrotaId=frota.id;dragEl=el;
    setTimeout(()=>el.classList.add('dragging'),0);
    e.dataTransfer.effectAllowed='move';
    ghostEl=el.cloneNode(true);
    ghostEl.className='card drag-ghost';
    ghostEl.style.cssText+=`;left:${{e.clientX}}px;top:${{e.clientY}}px`;
    document.body.appendChild(ghostEl);
    e.dataTransfer.setDragImage(new Image(),0,0);
  }});
  el.addEventListener('drag',e=>{{if(ghostEl&&e.clientX>0){{ghostEl.style.left=(e.clientX+12)+'px';ghostEl.style.top=(e.clientY-20)+'px';}}}});
  el.addEventListener('dragend',()=>{{
    el.classList.remove('dragging');
    if(ghostEl){{ghostEl.remove();ghostEl=null;}}
    dragFrotaId=dragEl=null;
  }});
  return el;
}}

function setupDrop(listEl,ccNome){{
  if(!PODE_EDITAR) return;
  listEl.addEventListener('dragover',e=>{{e.preventDefault();listEl.classList.add('drag-over');}});
  listEl.addEventListener('dragleave',e=>{{if(!listEl.contains(e.relatedTarget))listEl.classList.remove('drag-over');}});
  listEl.addEventListener('drop',e=>{{
    e.preventDefault();listEl.classList.remove('drag-over');
    if(!dragFrotaId)return;
    let ccAtual='';
    for(const cc of DATA.ccs){{if((DATA.frotas_por_cc[cc.nome]||[]).some(f=>f.id===dragFrotaId)){{ccAtual=cc.nome;break;}}}}
    if(ccAtual===ccNome)return;
    for(const cc of DATA.ccs){{
      const lista=DATA.frotas_por_cc[cc.nome];if(!lista)continue;
      const idx=lista.findIndex(f=>f.id===dragFrotaId);
      if(idx!==-1){{const[frota]=lista.splice(idx,1);frota.frente='';if(!DATA.frotas_por_cc[ccNome])DATA.frotas_por_cc[ccNome]=[];DATA.frotas_por_cc[ccNome].push(frota);break;}}
    }}
    render();
    const url=new URL(window.parent.location.href);
    url.searchParams.set('acao','mover_cc');
    url.searchParams.set('frota_id',dragFrotaId);
    url.searchParams.set('valor',ccNome==='📦 Sem alocação'?'':ccNome);
    window.parent.location.href=url.toString();
  }});
}}

function abrirModal(e,frotaId,frotaNome,frenteAtual){{
  e.stopPropagation();
  modalFrotaId=frotaId;
  document.getElementById('modalNome').textContent=frotaNome;
  document.getElementById('modalSub').textContent=frenteAtual?`Frente atual: ${{frenteAtual}}`:'Sem frente definida';
  const sel=document.getElementById('modalSelect');
  sel.innerHTML='';
  nomesFr.forEach(fr=>{{
    const opt=document.createElement('option');
    opt.value=fr==='— Sem frente —'?'__sem__':fr;
    opt.textContent=fr;
    if(fr===frenteAtual||(fr==='— Sem frente —'&&!frenteAtual))opt.selected=true;
    sel.appendChild(opt);
  }});
  document.getElementById('modalBg').classList.add('open');
}}
function fecharModal(){{document.getElementById('modalBg').classList.remove('open');modalFrotaId=null;}}
function salvarFrente(){{
  if(!modalFrotaId)return;
  const val=document.getElementById('modalSelect').value;
  fecharModal();
  const url=new URL(window.parent.location.href);
  url.searchParams.set('acao','mover_frente');
  url.searchParams.set('frota_id',modalFrotaId);
  url.searchParams.set('valor',val);
  window.parent.location.href=url.toString();
}}
function salvarObs(frotaId,texto){{
  const url=new URL(window.parent.location.href);
  url.searchParams.set('acao','salvar_obs');
  url.searchParams.set('frota_id',frotaId);
  url.searchParams.set('valor',texto);
  window.parent.location.href=url.toString();
}}
function marcarVendido(){{
  if(!modalFrotaId)return;
  fecharModal();
  const url=new URL(window.parent.location.href);
  url.searchParams.set('acao','vender');
  url.searchParams.set('frota_id',modalFrotaId);
  url.searchParams.set('valor','vendido');
  window.parent.location.href=url.toString();
}}
document.getElementById('modalBg').addEventListener('click',e=>{{if(e.target===document.getElementById('modalBg'))fecharModal();}});
render();
</script></body></html>"""

    # Componente bidirecional: renderiza o board e devolve as ações do usuário
    evento = kanban_board(data=board_data, key="board_kanban", default=None)

    if evento and isinstance(evento, dict) and PODE_EDITAR:
        if st.session_state.get("_board_nonce") != evento.get("nonce"):
            st.session_state["_board_nonce"] = evento.get("nonce")
            _ac  = evento.get("acao", "")
            _fid = str(evento.get("frota_id", "") or "")
            _val = evento.get("valor", "") or ""
            try:
                if _ac == "mover_cc" and _fid:
                    mover_cc(_fid, _val); st.rerun()
                elif _ac == "mover_frente" and _fid:
                    mover_frente(_fid, "" if _val == "__sem__" else _val); st.rerun()
                elif _ac == "salvar_obs" and _fid:
                    patch_item(LISTA_FROTAS, _fid, {"Obs": _val})
                    invalidar(); st.rerun()
                elif _ac == "vender" and _fid:
                    st.session_state["vender_frota_id"] = _fid
            except Exception as e:
                st.error(f"Erro ao executar ação: {e}")

    # ── Modal vender (via Streamlit) ──────────
    frota_id_venda = st.session_state.get("vender_frota_id", "")
    if frota_id_venda and PODE_EDITAR:
        frota_obj = next((f for f in frotas_ativas if f["id"] == frota_id_venda), None)
        if frota_obj:
            with st.form("form_venda"):
                st.subheader(f"🪦 Marcar como vendido: {frota_obj['nome']}")
                dv = st.date_input("Data da venda", value=datetime.today())
                vv = st.text_input("Valor de venda (R$)", placeholder="Ex: 150000")
                ov = st.text_input("Observação", placeholder="Comprador, motivo...")
                c1, c2 = st.columns(2)
                with c1:
                    if st.form_submit_button("✅ Confirmar venda", type="primary"):
                        vender_frota(frota_id_venda, str(dv), vv, ov)
                        st.session_state.pop("vender_frota_id", None); st.rerun()
                with c2:
                    if st.form_submit_button("Cancelar"):
                        st.session_state.pop("vender_frota_id", None); st.rerun()

# ══════════════════════════════════════════════
# ABA 2 — NOVA FROTA
# ══════════════════════════════════════════════
with aba_nova:
    st.subheader("➕ Cadastrar nova frota")
    if not PODE_EDITAR:
        st.info("👁️ Seu perfil é somente visualização — apenas administradores podem cadastrar frotas.")
    else:
        st.caption("Preencha os dados do novo equipamento adquirido.")

        with st.form("form_nova_frota", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                nn = st.text_input("Nome / Identificação *", placeholder="1600 - TRATOR NEW HOLLAND T7.245")
                nt = st.selectbox("Tipo *", TIPOS)
                nm = st.text_input("Modelo", placeholder="New Holland T7.245")
            with c2:
                nch = st.text_input("Chassi", placeholder="HCCZ7245XXXX")
                nan = st.text_input("Ano", placeholder="2026")
                nobs = st.text_input("Observação", placeholder="Notas sobre o equipamento")

            st.divider()
            st.markdown("**Alocação inicial (opcional)**")
            ca1, ca2 = st.columns(2)
            with ca1:
                ncc_opc = ["— Sem alocação (disponível) —"] + [c["nome"] for c in ccs]
                ncc = st.selectbox("Centro de Custo", ncc_opc)
            with ca2:
                nfr_opc = ["— Sem frente —"] + [fr["nome"] for fr in frentes]
                nfr = st.selectbox("Frente de Corte", nfr_opc)

            submitted = st.form_submit_button("💾 Cadastrar frota", type="primary", use_container_width=True)

        if submitted:
            if not nn.strip():
                st.error("Nome é obrigatório.")
            else:
                cc_val = "" if ncc == "— Sem alocação (disponível) —" else ncc
                fr_val = "" if nfr == "— Sem frente —" else nfr
                try:
                    criar_item(LISTA_FROTAS, {
                        "Title": nn.strip(), "Tipo": nt, "Chassi": nch.strip(),
                        "Ano": nan.strip(), "Obs": nobs.strip(),
                        "CCNome": cc_val, "FrenteNome": fr_val,
                        "Status": "Ativo", "DataVenda": "", "ValorVenda": "",
                    })
                    invalidar()
                    st.success(f"✅ Frota '{nn}' cadastrada com sucesso!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao cadastrar: {e}")

    # Lista de frotas disponíveis (sem CC)
    disponiveis = [f for f in frotas_ativas if not f["cc_nome"]]
    if disponiveis:
        st.divider()
        st.markdown(f"**📦 Frotas disponíveis sem alocação ({len(disponiveis)})**")
        for f in sorted(disponiveis, key=lambda x: x["nome"]):
            bg, tx = TIPO_COR.get(f["tipo"], ["#f3f4f6","#374151"])
            st.markdown(
                f'<div style="background:#fff;border:1px solid #e5e7eb;border-radius:8px;'
                f'padding:8px 12px;margin-bottom:6px;display:flex;align-items:center;gap:8px;">'
                f'<span style="font-size:13px;font-weight:600">{TIPO_EMOJI.get(f["tipo"],"📦")} {f["nome"]}</span>'
                f'<span style="background:{bg};color:{tx};padding:1px 8px;border-radius:20px;font-size:11px;font-weight:600">{f["tipo"]}</span>'
                f'<span style="font-size:11px;color:#6b7280">{f.get("chassi","")} {f.get("ano","")}</span>'
                f'</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════
# ABA 3 — ENTREGA FUTURA
# ══════════════════════════════════════════════
# Módulo à parte: registra frotas que HOJE estão alocadas em algum CC/frente
# mas que no futuro serão entregues/devolvidas (fim de contrato de aluguel,
# negociação de compra, liberação após operação, etc.). É só um aviso —
# nunca mexe no CCNome/FrenteNome da frota no board principal.
with aba_entrega:
    st.subheader("🚚 Entrega Futura")
    st.caption("Aviso de equipamentos que estão alocados hoje, mas que serão "
               "entregues/devolvidos futuramente. Não altera a posição da frota no board.")

    sub_pend, sub_hist = st.tabs([
        f"📋 Pendentes ({len(entregas_pendentes)})",
        f"✅ Histórico de entregas ({len(entregas_concluidas)})",
    ])

    # ── PENDENTES ──────────────────────────────
    with sub_pend:
        if PODE_EDITAR:
            with st.expander("➕ Registrar nova entrega futura", expanded=len(entregas_pendentes) == 0):
                with st.form("form_entrega_futura", clear_on_submit=True):
                    # Só oferece frotas que ainda não têm entrega pendente cadastrada
                    frotas_elegiveis = sorted(
                        [f for f in frotas_ativas if f["id"] not in ids_frota_com_entrega_pendente],
                        key=lambda x: x["nome"])
                    opcoes = [f'{f["nome"]}  ·  {f["cc_nome"] or "sem alocação"}' for f in frotas_elegiveis]

                    if not frotas_elegiveis:
                        st.info("Todas as frotas ativas já têm uma entrega futura pendente cadastrada.")
                        sel_idx = None
                    else:
                        sel_label = st.selectbox("Frota *", opcoes,
                            help="Busque pelo nome/número — os dados (tipo, chassi, CC/frente atuais) são puxados automaticamente")
                        sel_idx = opcoes.index(sel_label) if sel_label else None

                    ce1, ce2 = st.columns(2)
                    with ce1:
                        ne_motivo = st.text_area("Motivo / condição da entrega *",
                            placeholder="Ex: Contrato de aluguel encerra em dez/2026, será devolvida à New Agro")
                    with ce2:
                        ne_destino = st.text_input("Destino",
                            placeholder="Ex: New Agro, devolução ao proprietário...")

                    submitted_e = st.form_submit_button("💾 Registrar", type="primary", use_container_width=True)

                if submitted_e:
                    if sel_idx is None:
                        st.error("Selecione uma frota.")
                    elif not ne_motivo.strip():
                        st.error("Motivo é obrigatório.")
                    else:
                        frota_sel = frotas_elegiveis[sel_idx]
                        try:
                            registrar_entrega_futura(frota_sel, ne_motivo.strip(), ne_destino.strip())
                            st.success(f"✅ Entrega futura registrada para '{frota_sel['nome']}'!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erro ao registrar: {e}")

        st.divider()

        if not entregas_pendentes:
            st.info("Nenhuma entrega futura pendente no momento.")
        else:
            # Mapa rápido para achar a alocação ATUAL da frota (pode ter mudado
            # desde o cadastro, já que o board segue independente)
            frotas_por_id = {f["id"]: f for f in frotas_ativas}

            busca_e = st.text_input("🔍 Buscar", placeholder="Nome...", key="busca_e")
            pend_vis = [e for e in entregas_pendentes
                        if not busca_e or busca_e.lower() in e["nome"].lower()]
            st.markdown(f"*Exibindo {len(pend_vis)} de {len(entregas_pendentes)}*")

            for e in sorted(pend_vis, key=lambda x: x["nome"]):
                frota_atual = frotas_por_id.get(e["frota_id"])
                cc_atual = frota_atual["cc_nome"] if frota_atual else e.get("cc_origem", "—")
                fr_atual = frota_atual["frente_nome"] if frota_atual else e.get("frente_origem", "")
                mudou = frota_atual and frota_atual["cc_nome"] != e.get("cc_origem", "")

                with st.container(border=True):
                    c1, c2 = st.columns([3, 1])
                    with c1:
                        st.markdown(f"**🚚 {e['nome']}**  ·  {e.get('tipo','—')}")
                        st.caption(f"Chassi: {e.get('chassi','—')}  |  Cadastrado em: {e.get('data_cadastro','—')}")
                        st.markdown(f"📍 CC atual: **{cc_atual or 'sem alocação'}**"
                                    + (f" · ⚡ {fr_atual}" if fr_atual else "")
                                    + (" ⚠️ *(mudou desde o cadastro)*" if mudou else ""))
                        st.markdown(f"**Motivo:** {e.get('motivo','—')}")
                        if e.get("destino"):
                            st.markdown(f"**Destino:** {e['destino']}")
                    with c2:
                        if PODE_EDITAR:
                            if st.button("✅ Marcar entregue", key=f"entregar_{e['id']}", use_container_width=True):
                                st.session_state["baixa_entrega_id"] = e["id"]
                            if st.button("🗑️ Cancelar registro", key=f"cancelar_{e['id']}", use_container_width=True):
                                cancelar_entrega_futura(e["id"])
                                st.rerun()

                    if PODE_EDITAR and st.session_state.get("baixa_entrega_id") == e["id"]:
                        with st.form(f"form_baixa_{e['id']}"):
                            st.markdown(f"**Confirmar entrega de {e['nome']}**")
                            bd1, bd2 = st.columns(2)
                            with bd1:
                                data_ent = st.date_input("Data da entrega", value=datetime.today())
                            with bd2:
                                obs_ent = st.text_input("Observação", placeholder="Opcional")
                            bc1, bc2 = st.columns(2)
                            with bc1:
                                if st.form_submit_button("✅ Confirmar", type="primary", use_container_width=True):
                                    marcar_entrega_concluida(e["id"], str(data_ent), obs_ent)
                                    del st.session_state["baixa_entrega_id"]
                                    st.success("Entrega confirmada!")
                                    st.rerun()
                            with bc2:
                                if st.form_submit_button("Cancelar", use_container_width=True):
                                    del st.session_state["baixa_entrega_id"]
                                    st.rerun()

    # ── HISTÓRICO ──────────────────────────────
    with sub_hist:
        if not entregas_concluidas:
            st.info("Nenhuma entrega concluída ainda.")
        else:
            for e in sorted(entregas_concluidas, key=lambda x: x.get("data_entrega",""), reverse=True):
                with st.expander(f"✅ {e['nome']} — entregue em {e.get('data_entrega','?')}"):
                    c1, c2 = st.columns(2)
                    c1.markdown(f"**Tipo:** {e.get('tipo','—')}")
                    c1.markdown(f"**Chassi:** {e.get('chassi','—')}")
                    c1.markdown(f"**CC de origem:** {e.get('cc_origem','—')}")
                    c2.markdown(f"**Motivo:** {e.get('motivo','—')}")
                    c2.markdown(f"**Destino:** {e.get('destino','—')}")
                    c2.markdown(f"**Obs. da baixa:** {e.get('obs_baixa','—')}")

# ══════════════════════════════════════════════
# ABA 4 — FROTAS VENDIDAS / DESCARTE
# ══════════════════════════════════════════════
with aba_vendidos:
    st.subheader(f"📋 Histórico de Baixas — {len(frotas_vendidas)} equipamentos")
    st.caption("Equipamentos vendidos, descartados ou baixados do patrimônio.")

    if not frotas_vendidas:
        st.info("Nenhuma frota vendida ou descartada ainda.")
    else:
        # Filtros
        bv1, bv2 = st.columns([2,2])
        with bv1:
            busca_v = st.text_input("🔍 Buscar", placeholder="Nome...", key="busca_v")
        with bv2:
            filtro_tipo_v = st.multiselect("Tipo", TIPOS, default=[], key="ft_v", placeholder="Todos")

        vendidos_vis = [f for f in frotas_vendidas
                        if (not busca_v or busca_v.lower() in f["nome"].lower())
                        and (not filtro_tipo_v or f["tipo"] in filtro_tipo_v)]

        st.markdown(f"*Exibindo {len(vendidos_vis)} de {len(frotas_vendidas)}*")
        st.divider()

        for f in sorted(vendidos_vis, key=lambda x: x.get("data_venda",""), reverse=True):
            bg, tx = TIPO_COR.get(f["tipo"], ["#f3f4f6","#374151"])
            with st.expander(f"🪦 {f['nome']} — vendido em {f.get('data_venda','?')}"):
                c1,c2,c3 = st.columns(3)
                c1.markdown(f"**Tipo:** {f['tipo']}")
                c1.markdown(f"**Chassi:** {f.get('chassi','—')}")
                c2.markdown(f"**Ano:** {f.get('ano','—')}")
                c2.markdown(f"**Valor venda:** R$ {f.get('valor_venda','—')}")
                c3.markdown(f"**Data venda:** {f.get('data_venda','—')}")
                c3.markdown(f"**Obs:** {f.get('obs','—')}")
                if PODE_EDITAR and st.button("↩️ Reativar frota", key=f"reativ_{f['id']}"):
                    reativar_frota(f["id"])
                    st.success("Frota reativada!"); st.rerun()

# ══════════════════════════════════════════════
# ABA 5 — EXPORTAR RELATÓRIO
# ══════════════════════════════════════════════
with aba_export:
    st.subheader("📤 Exportar relatório")

    ex1, ex2, ex3 = st.columns([2,1,1])
    with ex1: mes_ref = st.text_input("Mês de referência", value=datetime.now().strftime("%B/%Y"))
    with ex2: safra   = st.text_input("Safra", value="2025/26")
    with ex3:
        filtro_export_cc = st.selectbox(
            "Filtrar por CC",
            ["Todos os CCs"] + [c["nome"] for c in ccs],
            help="Exporta só as frotas deste CC, ou todos")

    st.markdown("**Selecione as abas a incluir:**")
    ea, eb, ec, ed = st.columns(4)
    inc_det  = ea.checkbox("Alocação detalhada", value=True)
    inc_cc   = eb.checkbox("Resumo por CC",       value=True)
    inc_fr   = ec.checkbox("Por frente",          value=True)
    inc_vend = ed.checkbox("Vendidos/Descarte",   value=False)

    if st.button("⬇️ Gerar Excel", type="primary", use_container_width=True):

        # Frotas a exportar
        frotas_exp = frotas_ativas if filtro_export_cc == "Todos os CCs" else \
                     [f for f in frotas_ativas if f["cc_nome"] == filtro_export_cc]

        wb_e = Workbook(); wb_e.remove(wb_e.active)

        def hdr_s(ws, row, cols, bg, fg="FFFFFF"):
            fill = PatternFill("solid", fgColor=bg.lstrip("#"))
            font = Font(bold=True, color=fg, size=10)
            aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
            bdr  = Border(left=Side("thin","CCCCCC"), right=Side("thin","CCCCCC"),
                          top=Side("thin","CCCCCC"),  bottom=Side("thin","CCCCCC"))
            for c in cols:
                cell = ws.cell(row=row, column=c)
                cell.fill=fill; cell.font=font; cell.alignment=aln; cell.border=bdr

        def lin_s(ws, row, cols, bg="FFFFFF"):
            fill = PatternFill("solid", fgColor=bg.lstrip("#"))
            bdr  = Border(left=Side("thin","E5E7EB"), right=Side("thin","E5E7EB"),
                          top=Side("thin","E5E7EB"),  bottom=Side("thin","E5E7EB"))
            for c in cols:
                ws.cell(row=row,column=c).fill=fill
                ws.cell(row=row,column=c).border=bdr
                ws.cell(row=row,column=c).font=Font(size=10)

        titulo_base = f"FROTAS — {filtro_export_cc.upper() if filtro_export_cc != 'Todos os CCs' else 'TODOS OS CCs'} | {mes_ref.upper()} | SAFRA {safra}"

        if inc_det:
            ws1 = wb_e.create_sheet("Alocação Detalhada")
            ws1.merge_cells("A1:H1"); ws1["A1"] = titulo_base
            ws1["A1"].font=Font(bold=True,size=12,color="FFFFFF")
            ws1["A1"].fill=PatternFill("solid",fgColor="1D9E75")
            ws1["A1"].alignment=Alignment(horizontal="center",vertical="center")
            ws1.row_dimensions[1].height=26
            ws1.merge_cells("A2:H2")
            ws1["A2"]=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws1["A2"].font=Font(italic=True,size=9,color="6B7280")
            ws1["A2"].alignment=Alignment(horizontal="right")
            hdrs1=["Frota","Tipo","Chassi","Ano","Centro de Custo","Frente de Corte","Observação","Situação"]
            for ci,h in enumerate(hdrs1,1): ws1.cell(row=3,column=ci,value=h)
            hdr_s(ws1,3,range(1,9),"0F6E56"); ws1.row_dimensions[3].height=18
            row=4
            for f in sorted(frotas_exp, key=lambda x:(x.get("cc_nome","zzz"),x.get("frente_nome","zzz"),x["nome"])):
                sit="✅ Alocada" if f["cc_nome"] else "⚠️ Disponível"
                linha=[f["nome"],f["tipo"],f.get("chassi",""),f.get("ano",""),
                       f["cc_nome"] or "—",f["frente_nome"] or "—",f.get("obs",""),sit]
                for ci,v in enumerate(linha,1): ws1.cell(row=row,column=ci,value=v)
                lin_s(ws1,row,range(1,9),"F0FDF4" if f["cc_nome"] else "FFF7ED"); row+=1
            for ci,w in zip(range(1,9),[32,14,22,8,30,24,22,14]):
                ws1.column_dimensions[get_column_letter(ci)].width=w
            ws1.freeze_panes="A4"

        if inc_cc:
            ws2 = wb_e.create_sheet("Resumo por CC")
            ws2.merge_cells("A1:H1"); ws2["A1"]=f"RESUMO POR CC — {mes_ref.upper()}"
            ws2["A1"].font=Font(bold=True,size=12,color="FFFFFF")
            ws2["A1"].fill=PatternFill("solid",fgColor="378ADD")
            ws2["A1"].alignment=Alignment(horizontal="center",vertical="center")
            ws2.row_dimensions[1].height=26
            hdrs2=["Centro de Custo","Total","Colhedoras","Transbordos","Tratores","Caminhões","Veículos","Outros"]
            for ci,h in enumerate(hdrs2,1): ws2.cell(row=2,column=ci,value=h)
            hdr_s(ws2,2,range(1,9),"185FA5"); r2=3
            ccs_rel = [c for c in ccs if filtro_export_cc=="Todos os CCs" or c["nome"]==filtro_export_cc]
            for cc in ccs_rel:
                fc=[f for f in frotas_ativas if f["cc_nome"]==cc["nome"]]
                linha2=[cc["nome"],len(fc),
                        sum(1 for f in fc if f["tipo"]=="Colhedora"),
                        sum(1 for f in fc if f["tipo"]=="Transbordo"),
                        sum(1 for f in fc if f["tipo"]=="Trator"),
                        sum(1 for f in fc if f["tipo"]=="Caminhão"),
                        sum(1 for f in fc if f["tipo"]=="Veículo"),
                        sum(1 for f in fc if f["tipo"] in("Implemento","Apoio","Área de Vivência","Outro"))]
                for ci,v in enumerate(linha2,1): ws2.cell(row=r2,column=ci,value=v)
                lin_s(ws2,r2,range(1,9),"EFF6FF" if r2%2==0 else "FFFFFF"); r2+=1
            ws2.cell(r2,1,"TOTAL")
            for ci in range(2,9): ws2.cell(r2,ci,f"=SUM({get_column_letter(ci)}3:{get_column_letter(ci)}{r2-1})")
            hdr_s(ws2,r2,range(1,9),"1D9E75")
            for ci,w in zip(range(1,9),[34,8,12,12,10,12,10,10]):
                ws2.column_dimensions[get_column_letter(ci)].width=w

        if inc_fr:
            ws3 = wb_e.create_sheet("Por Frente")
            ws3.merge_cells("A1:G1"); ws3["A1"]=f"POR FRENTE — {mes_ref.upper()}"
            ws3["A1"].font=Font(bold=True,size=12,color="FFFFFF")
            ws3["A1"].fill=PatternFill("solid",fgColor="BA7517")
            ws3["A1"].alignment=Alignment(horizontal="center",vertical="center")
            ws3.row_dimensions[1].height=26
            hdrs3=["Frente","Centro de Custo","Frota","Tipo","Chassi","Ano","Observação"]
            for ci,h in enumerate(hdrs3,1): ws3.cell(row=2,column=ci,value=h)
            hdr_s(ws3,2,range(1,8),"854F0B"); r3=3
            for fr in frentes:
                ff=[f for f in frotas_exp if f["frente_nome"]==fr["nome"]]
                for f in sorted(ff,key=lambda x:x["nome"]):
                    linha3=[fr["nome"],f["cc_nome"] or "—",f["nome"],f["tipo"],
                            f.get("chassi",""),f.get("ano",""),f.get("obs","")]
                    for ci,v in enumerate(linha3,1): ws3.cell(row=r3,column=ci,value=v)
                    lin_s(ws3,r3,range(1,8),"FFFBEB" if r3%2==0 else "FFFFFF"); r3+=1
            for ci,w in zip(range(1,8),[24,30,34,14,22,8,22]):
                ws3.column_dimensions[get_column_letter(ci)].width=w

        if inc_vend:
            ws4 = wb_e.create_sheet("Vendidos-Descarte")
            ws4.merge_cells("A1:G1"); ws4["A1"]="FROTAS VENDIDAS / DESCARTE"
            ws4["A1"].font=Font(bold=True,size=12,color="FFFFFF")
            ws4["A1"].fill=PatternFill("solid",fgColor="A32D2D")
            ws4["A1"].alignment=Alignment(horizontal="center",vertical="center")
            ws4.row_dimensions[1].height=26
            hdrs4=["Frota","Tipo","Chassi","Ano","Data Venda","Valor Venda (R$)","Observação"]
            for ci,h in enumerate(hdrs4,1): ws4.cell(row=2,column=ci,value=h)
            hdr_s(ws4,2,range(1,8),"791F1F"); r4=3
            for f in sorted(frotas_vendidas, key=lambda x:x.get("data_venda",""), reverse=True):
                linha4=[f["nome"],f["tipo"],f.get("chassi",""),f.get("ano",""),
                        f.get("data_venda",""),f.get("valor_venda",""),f.get("obs","")]
                for ci,v in enumerate(linha4,1): ws4.cell(row=r4,column=ci,value=v)
                lin_s(ws4,r4,range(1,8),"FEF2F2" if r4%2==0 else "FFFFFF"); r4+=1
            for ci,w in zip(range(1,8),[32,14,22,8,14,16,24]):
                ws4.column_dimensions[get_column_letter(ci)].width=w

        buf=io.BytesIO(); wb_e.save(buf); buf.seek(0)
        cc_slug = filtro_export_cc.replace(" ","_").replace("/","-") if filtro_export_cc!="Todos os CCs" else "todos"
        nome_arq=f"frotas_{cc_slug}_{datetime.now().strftime('%Y%m_%B').lower()}.xlsx"
        st.download_button("📥 Baixar Excel", buf, nome_arq,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, type="primary")
        n_abas = sum([inc_det,inc_cc,inc_fr,inc_vend])
        st.success(f"Relatório gerado com {n_abas} aba(s) — {len(frotas_exp)} frotas.")